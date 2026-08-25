from __future__ import annotations

from pathlib import Path

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader


def extract_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".txt", ".md", ".py", ".js", ".ts", ".json", ".csv", ".yaml", ".yml"}:
        return path.read_text(encoding="utf-8", errors="replace")[:200_000]
    if suffix == ".pdf":
        return "\n".join(page.extract_text() or "" for page in PdfReader(str(path)).pages)[:200_000]
    if suffix == ".docx":
        doc = Document(str(path))
        return "\n".join(p.text for p in doc.paragraphs)[:200_000]
    if suffix == ".xlsx":
        wb = load_workbook(path, read_only=True, data_only=True)
        rows: list[str] = []
        for ws in wb.worksheets:
            rows.append(f"## {ws.title}")
            for row in ws.iter_rows(values_only=True):
                rows.append(" | ".join("" if value is None else str(value) for value in row))
                if sum(map(len, rows)) > 200_000:
                    break
        return "\n".join(rows)[:200_000]
    return "Unsupported file type."
